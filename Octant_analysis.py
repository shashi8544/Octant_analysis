

###### define mod value
mod=5000


##### importing libraries
try:
    from datetime import datetime
    start_time = datetime.now()
    import os
    import glob
    import pandas as pd
    from cmath import nan
    import openpyxl
    from openpyxl.styles import Border, Side
    from openpyxl.styles import PatternFill
except:
    print("there is some error in importing libraries check whether all libraries is installed or not")
    exit()

#### checking python version
try:
    from platform import python_version
    ver = python_version()

    if ver == "3.8.10":
        print("Correct Version Installed")
    else:
        print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")
except:
    print("please install platform libraries")


######## Reading excel input file
try:
    os.chdir(r"C:\Users\DELL\OneDrive\Desktop\tt\tut07\input")
    extension='xlsx'
    all_filesname = [i for i in glob.glob('*.{}'.format(extension))]
except:
    print("there is some error in reading excel file!!!!!!!!! Please check ")

##### tutorial 5 defnition
def tut5(wb,mod):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    #### calculating average value
    try:
        sheet = wb.active
        sheet.column_dimensions['O'].width=40
        sheet.column_dimensions['AE'].width=40
        Uavg=0
        Vavg=0
        Wavg=0
        ls=[]
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=19888, max_col=4):
            lst1=[]
            for cell in row:
                lst1.append(cell.value)
            ls.append(lst1)
        for i in range(1,19888):
            Uavg=Uavg+ls[i][1]
            Vavg=Vavg+ls[i][2]
            Wavg=Wavg+ls[i][3]
        Uavg=Uavg/19887
        Vavg=Vavg/19887
        Wavg=Wavg/19887

        lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
    except:
        print("there is error in calculating average value")
        exit()

    ### writing nan in tansition value

    try:
        i=0
        for row in sheet.iter_rows(min_row=1, min_col=12, max_row=29777, max_col=51):
            j=0
            for cell in row:
                cell.value=nan
                j=j+1
            i=i+1
    except:
        print("there is some error in transition count")
    

    #####creating octant of given value

    try:
        i=0
        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
            j=0
            for cell in row:
                cell.value=lst_avg[i][j]
                j=j+1
            i=i+1
        lst_newval=[]
        for i in range(1,19888):
            lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])
        sheet.cell(row=1,column=8).value="Uavg'"
        sheet.cell(row=1,column=9).value="Vavg'"
        sheet.cell(row=1,column=10).value="Wavg'"    
        sheet.cell(row=1,column=11).value="Octant"    

        i=0
        for row in sheet.iter_rows(min_row=2, min_col=8, max_row=19888, max_col=10):
            j=0
            for cell in row:
                cell.value=lst_newval[i][j]
                j=j+1
            i=i+1

        lst_octant = []
        for p in lst_newval:
            if(p[0]>=0):
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(1)
                    else:
                        lst_octant.append(-1)
                else:
                    if(p[2]>=0):
                        lst_octant.append(4)
                    else:
                        lst_octant.append(-4)
            else:
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(2)
                    else:
                        lst_octant.append(-2)
                else:
                    if(p[2]>=0):
                        lst_octant.append(3)
                    else:
                        lst_octant.append(-3)

        i=0
        for row in sheet.iter_rows(min_row=2, min_col=11, max_row=19888, max_col=11):
            for cell in row:
                cell.value=lst_octant[i]
            i=i+1
    except:
        print("there is error in creating octant value")
        exit()
    
    ###### creating overall count

    try:
        t=0
        if(19888%mod==0):
            t=19888//mod
        else:
            t=(19888//mod)+1

        sheet.cell(row=2,column=14).value=1
        sheet.cell(row=2,column=15).value=-1
        sheet.cell(row=2,column=16).value=2
        sheet.cell(row=2,column=17).value=-2
        sheet.cell(row=2,column=18).value=3
        sheet.cell(row=2,column=19).value=-3
        sheet.cell(row=2,column=20).value=4
        sheet.cell(row=2,column=21).value=-4
        sheet.cell(row=3,column=13).value="Overall Count"
        sheet.cell(row=4,column=12).value="User Input"
        tt=str(mod)
        sheet.cell(row=4,column=13).value="mod"+" "+tt
        lst_overall_count = [0,0,0,0,0,0,0,0]
        for valu in lst_octant:
            if valu==1:
                lst_overall_count[0]=lst_overall_count[0]+1
            if valu==-1:
                lst_overall_count[1]=lst_overall_count[1]+1
            if valu==2:
                lst_overall_count[2]=lst_overall_count[2]+1
            if valu==-2:
                lst_overall_count[3]=lst_overall_count[3]+1
            if valu==3:
                lst_overall_count[4]=lst_overall_count[4]+1
            if valu==-3:
                lst_overall_count[5]=lst_overall_count[5]+1
            if valu==4:
                lst_overall_count[6]=lst_overall_count[6]+1
            if valu==-4:
                lst_overall_count[7]=lst_overall_count[7]+1
        for row in sheet.iter_rows(min_row=3, min_col=14, max_row=3, max_col=21):
            j=0
            for cell in row:
                cell.value=lst_overall_count[j]
                j=j+1
        lst_hh=[]
        
        for j in range(t):
            if(j==t-1):
                ttm1=str(mod*j)
                ttm=ttm1+"-"+"19887"
                lst_hh.append(ttm)
            else:
                if(j==0):
                    ttm1=".0000"
                else:
                    ttm1=str(mod*j) 
                ttm2=str(mod*(j+1)-1)
                ttm=ttm1+"-"+ttm2
                lst_hh.append(ttm) 

    except:
        print("there is error in creating overall count value")
        exit()

    ##### updating overall count and transition
    try:
        lst_hh_val=[]
        for j in range(t):
            lst_hh_temp=[0,0,0,0,0,0,0,0]
            if(j==t-1):
                y=19887
            else:
                y=mod*(j+1)
            for valu in range(mod*j,y):
                if lst_octant[valu]==1:
                    lst_hh_temp[0]=lst_hh_temp[0]+1
                if lst_octant[valu]==-1:
                    lst_hh_temp[1]=lst_hh_temp[1]+1
                if lst_octant[valu]==2:
                    lst_hh_temp[2]=lst_hh_temp[2]+1
                if lst_octant[valu]==-2:
                    lst_hh_temp[3]=lst_hh_temp[3]+1
                if lst_octant[valu]==3:
                    lst_hh_temp[4]=lst_hh_temp[4]+1
                if lst_octant[valu]==-3:
                    lst_hh_temp[5]=lst_hh_temp[5]+1
                if lst_octant[valu]==4:
                    lst_hh_temp[6]=lst_hh_temp[6]+1
                if lst_octant[valu]==-4:
                    lst_hh_temp[7]=lst_hh_temp[7]+1


            lst_hh_val.append(lst_hh_temp)
        for i in range(t):
            sheet.cell(row=i+5,column=13).value=lst_hh[i]
        for i in range(t):
            for j in range(8):
                sheet.cell(row=i+5,column=14+j).value=lst_hh_val[i][j]

    except:
        print("there is error in updating in verified count")
        
    #### creating skelton and list of rank
    try:
        lst_rec=[1,-1,2,-2,3,-3,4,-4]
        lst_rec1=["1","-1","2","-2","3","-3","4","-4"]
        lst_rank_head=["Rank 1","Rank 2","Rank 3","Rank 4","Rank 5","Rank 6","Rank 7","Rank 8","Rank1 OctantID","Rank1 Octant Name"]
        for i in range(len(lst_rec)):
            sheet.cell(row=1,column=i+22).value=lst_rec[i]
        for i in range(len(lst_rank_head)):
            sheet.cell(row=2,column=i+22).value=lst_rank_head[i]
        lst_record=[]
        lst_overall_count1=lst_overall_count.copy()
        lst_overall_count1.sort(reverse=True)
        list_temp=[]
        for i in range(8):
            for j in range(8):
                if(lst_overall_count[i]==lst_overall_count1[j]):
                    if(j==0):
                        lst_record.append(i)
                    list_temp.append(j+1)
                    break
        for i in range(8):
            sheet.cell(row=3,column=i+22).value=list_temp[i]
        sheet.cell(row=3,column=30).value=lst_record[0]
        t1=str(lst_rec[lst_record[0]])
        sheet.cell(row=3,column=31).value=octant_name_id_mapping[t1]

        lst_rank=[]
        lst_record=[]
        for i in range(t):
            lst_hh_val1=lst_hh_val[i].copy()
            lst_hh_val1.sort(reverse=True)
            list_temp=[]
            for k in range(8):
                for j in range(8):
                    if(lst_hh_val[i][k]==lst_hh_val1[j]):
                        if(j==0):
                            lst_record.append(k)
                        list_temp.append(j+1)
                        break
            lst_rank.append(list_temp)
        for i in range(t):
            for j in range(8):
                sheet.cell(row=i+5,column=j+22).value=lst_rank[i][j]
                if(sheet.cell(row=i+5,column=j+22).value==1):
                    sheet.cell(row=i+5,column=j+22).fill=PatternFill(patternType='solid',fgColor="FCBA03")
        lst_record1=[]
        lst_record2=[]
        for item in lst_record:
            lst_record1.append(str(lst_rec[item]))
            lst_record2.append((lst_rec[item]))
        for i in range(len(lst_record2)):
            sheet.cell(row=i+5,column=30).value=lst_record2[i]
            sheet.cell(row=i+5,column=31).value=octant_name_id_mapping[lst_record1[i]]
        lst_hh1=["Octant ID","Octant Name","Count of Rank1 Mod Value"]
        for i in range(3):
            sheet.cell(row=t+8,column=i+29).value=lst_hh1[i]
        for i in range(8):
            sheet.cell(row=i+t+9,column=29).value=lst_rec[i]
            sheet.cell(row=i+t+9,column=30).value=octant_name_id_mapping[lst_rec1[i]]
        dict_rec={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
        for item in lst_record2:
            dict_rec[item]=dict_rec[item]+1
        
        for i in range(8):
            sheet.cell(row=i+t+9,column=31).value=dict_rec[lst_rec[i]]
        
    except:
        print("there is some errror in creating skelton and list of rank")


#### tutorial 4 definition
def tut4(wb,mod):
        #### calculating average value
    sheet = wb.active
    try:
        Uavg=0
        Vavg=0
        Wavg=0
        ls=[]
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=19888, max_col=4):
            lst1=[]
            for cell in row:
                lst1.append(cell.value)
            ls.append(lst1)
        for i in range(1,19888):
            Uavg=Uavg+ls[i][1]
            Vavg=Vavg+ls[i][2]
            Wavg=Wavg+ls[i][3]
        Uavg=Uavg/19887
        Vavg=Vavg/19887
        Wavg=Wavg/19887

        lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
    except:
        print("there is error in calculating average value")
        exit()

    #####creating octant of given value

    try:
        i=0
        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
            j=0
            for cell in row:
                cell.value=lst_avg[i][j]
                j=j+1
            i=i+1
        lst_newval=[]
        for i in range(1,19888):
            lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])   

        lst_octant = []
        for p in lst_newval:
            if(p[0]>=0):
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(1)
                    else:
                        lst_octant.append(-1)
                else:
                    if(p[2]>=0):
                        lst_octant.append(4)
                    else:
                        lst_octant.append(-4)
            else:
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(2)
                    else:
                        lst_octant.append(-2)
                else:
                    if(p[2]>=0):
                        lst_octant.append(3)
                    else:
                        lst_octant.append(-3)

    except:
        print("there is error in creating octant value")
        exit()



    ### creating skelton table
    try:
        row_head = ["count","Longest Subsquence Length","count"]
        row_column = ["+1","-1","+2","-2","+3","-3","+4","-4"]
        for row in sheet.iter_rows(min_row=2, min_col=45, max_row=2, max_col=47):
            j=0
            for cell in row:
                cell.value=row_head[j]
                j=j+1
        for row in sheet.iter_rows(min_row=2, min_col=49, max_row=2, max_col=51):
            j=0
            for cell in row:
                cell.value=row_head[j]
                j=j+1

        i=0
        for row in sheet.iter_rows(min_row=3, min_col=45, max_row=10, max_col=45):
            for cell in row:
                cell.value=row_column[i]
            i=i+1
    except:
        print("there is error in creating skelton of table")




    #####  subsequence codes 
    try:
        prev1=[1,-1,2,-2,3,-3,4,-4]
        ans=[]

        for prev in prev1:
            count1_max=0
            count1=0
            temp_count=0
            for v in lst_octant:
                if(v==prev):
                    temp_count=temp_count+1
                else:
                    if(temp_count>count1_max):
                        count1=1
                        count1_max=temp_count
                    elif(count1_max==temp_count & count1_max!=0):
                        count1=count1+1
                    temp_count=0
            lst_temp=[count1_max,count1]
            ans.append(lst_temp)
            count1_max=0
            count1=0
            temp_count=0
    except:
        print("there is some error in codes of subsequence")


    ##### updating subsequence count in excel
    try:
        i=0
        for row in sheet.iter_rows(min_row=3, min_col=46, max_row=10, max_col=47):
            j=0
            for cell in row:
                cell.value=ans[i][j]
                j=j+1
            i=i+1
    except:
        print("there is some error in updating excel file")

    #### finding timestamp of maximum subsequence
    res=[]
    count_line=0
    list_temp=["Time","From","To"]
    for i in range(8):
        list_temp_1=[]
        list_temp_1.append(row_column[i])
        list_temp_1.append(ans[i][0])
        list_temp_1.append(ans[i][1])
        res.append(list_temp_1)


    #########creating list of time stamp
    try:   
        new_time_list=[]
        x=0.0
        y=0.0
        for i in range(1,len(ls)):
            new_time_list.append(ls[i][0])
        time_list=[]
        temp=[1,-1,2,-2,3,-3,4,-4]
        for i in range(8):
            time_list_1=[]
            temp_count=0
            for j in range(19887):
                if(temp_count==ans[i][0]):
                    z=[x,y]
                    time_list_1.append(z)
                    x=0.0
                    y=0.0
                    temp_count=0
                if(temp[i]==lst_octant[j]):
                    if(temp_count==0):
                        x=new_time_list[j]
                        y=new_time_list[j]
                    else:
                        y=new_time_list[j]
                    temp_count=temp_count+1
                else:
                    x=0.0
                    y=0.0
                    temp_count=0
            time_list.append(time_list_1)
    except:
        print("Error in calculating time stamp")


    ########  updating time stamp in excel
    try:
        ro=3
        for i in range(8):
            for row in sheet.iter_rows(min_row=ro, min_col=49, max_row=ro, max_col=51):
                j=0
                for cell in row:
                    cell.value=res[i][j]
                    cell.border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
                    j=j+1
            ro=ro+1
            for row in sheet.iter_rows(min_row=ro, min_col=49, max_row=ro, max_col=51):
                j=0
                for cell in row:
                    cell.value=list_temp[j]
                    cell.border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
                    j=j+1
            ro=ro+1
            for items in time_list[i]:
                for row in sheet.iter_rows(min_row=ro, min_col=50, max_row=ro, max_col=51):
                    j=0
                    for cell in row:
                        cell.value=items[j]
                        cell.border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
                        cell.border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
                        j=j+1
                ro=ro+1

    except:
        print("there is error in uploading time stamp in excel")

#### tutorial 2 definition
def tut2(wb,mod):
        #### calculating average value
    try:
        sheet = wb.active
        Uavg=0
        Vavg=0
        Wavg=0
        ls=[]
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=19888, max_col=4):
            lst1=[]
            for cell in row:
                lst1.append(cell.value)
            ls.append(lst1)
        for i in range(1,19888):
            Uavg=Uavg+ls[i][1]
            Vavg=Vavg+ls[i][2]
            Wavg=Wavg+ls[i][3]
        Uavg=Uavg/19887
        Vavg=Vavg/19887
        Wavg=Wavg/19887

        lst_avg=[["Uavg","Vavg","Wavg"],[Uavg,Vavg,Wavg]]
    except:
        print("there is error in calculating average value")
        exit()
    #####creating octant of given value

    try:
        i=0
        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=2, max_col=7):
            j=0
            for cell in row:
                cell.value=lst_avg[i][j]
                j=j+1
            i=i+1
        lst_newval=[]
        for i in range(1,19888):
            lst_newval.append([ls[i][1]-Uavg,ls[i][2]-Vavg,ls[i][3]-Wavg])
        lst_octant = []
        for p in lst_newval:
            if(p[0]>=0):
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(1)
                    else:
                        lst_octant.append(-1)
                else:
                    if(p[2]>=0):
                        lst_octant.append(4)
                    else:
                        lst_octant.append(-4)
            else:
                if(p[1]>=0):
                    if(p[2]>=0):
                        lst_octant.append(2)
                    else:
                        lst_octant.append(-2)
                else:
                    if(p[2]>=0):
                        lst_octant.append(3)
                    else:
                        lst_octant.append(-3)
    except:
        print("there is error in creating octant value")
        exit()

    ###### there is some error in creating overall count

    try:
        t=0
        if(19888%mod==0):
            t=19888//mod
        else:
            t=(19888//mod)+1

        tt=str(mod)
        lst_overall_count = [0,0,0,0,0,0,0,0]
        for valu in lst_octant:
            if valu==1:
                lst_overall_count[0]=lst_overall_count[0]+1
            if valu==-1:
                lst_overall_count[1]=lst_overall_count[1]+1
            if valu==2:
                lst_overall_count[2]=lst_overall_count[2]+1
            if valu==-2:
                lst_overall_count[3]=lst_overall_count[3]+1
            if valu==3:
                lst_overall_count[4]=lst_overall_count[4]+1
            if valu==-3:
                lst_overall_count[5]=lst_overall_count[5]+1
            if valu==4:
                lst_overall_count[6]=lst_overall_count[6]+1
            if valu==-4:
                lst_overall_count[7]=lst_overall_count[7]+1
        lst_hh=[]
        for j in range(t):
            if(j==t-1):
                ttm1=str(mod*j)
                ttm=ttm1+"-"+"19887"
                lst_hh.append(ttm)
            else:
                if(j==0):
                    ttm1=".0000"
                else:
                    ttm1=str(mod*j) 
                ttm2=str(mod*(j+1)-1)
                ttm=ttm1+"-"+ttm2
                lst_hh.append(ttm) 
        lst_hh.append("Verified")
    except:
        print("there is error in creating overall count value and verifing it")
        exit()

    ##### updating overall count and transition
    try:
        lst_hh_val=[]
        for j in range(t):
            lst_hh_temp=[0,0,0,0,0,0,0,0]
            if(j==t-1):
                y=19887
            else:
                y=mod*(j+1)
            for valu in range(mod*j,y):
                if lst_octant[valu]==1:
                    lst_hh_temp[0]=lst_hh_temp[0]+1
                if lst_octant[valu]==-1:
                    lst_hh_temp[1]=lst_hh_temp[1]+1
                if lst_octant[valu]==2:
                    lst_hh_temp[2]=lst_hh_temp[2]+1
                if lst_octant[valu]==-2:
                    lst_hh_temp[3]=lst_hh_temp[3]+1
                if lst_octant[valu]==3:
                    lst_hh_temp[4]=lst_hh_temp[4]+1
                if lst_octant[valu]==-3:
                    lst_hh_temp[5]=lst_hh_temp[5]+1
                if lst_octant[valu]==4:
                    lst_hh_temp[6]=lst_hh_temp[6]+1
                if lst_octant[valu]==-4:
                    lst_hh_temp[7]=lst_hh_temp[7]+1


            lst_hh_val.append(lst_hh_temp)

        lst_verified=[0,0,0,0,0,0,0,0]
        for i in range(t):
            for j in range(8):
                lst_verified[j]=lst_verified[j]+lst_hh_val[i][j]

    except:
        print("there is error in updating in verified count")


    #### creating list for mod transition value
    try:
        lst_tran_count=[]

        for i in range(t):
            if(i==t-1):
                y=19886
            else:
                y=mod*(i+1)-1
            lst_tran_count_temp=[]
            temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            for k in range(mod*i,y):

                if(lst_octant[k]==1):
                    temp_1[lst_octant[k+1]]=temp_1[lst_octant[k+1]]+1
                if(lst_octant[k]==-1):
                    temp_11[lst_octant[k+1]]=temp_11[lst_octant[k+1]]+1
                if(lst_octant[k]==2):
                    temp_2[lst_octant[k+1]]=temp_2[lst_octant[k+1]]+1
                if(lst_octant[k]==-2):
                    temp_22[lst_octant[k+1]]=temp_22[lst_octant[k+1]]+1
                if(lst_octant[k]==3):
                    temp_3[lst_octant[k+1]]=temp_3[lst_octant[k+1]]+1
                if(lst_octant[k]==-3):
                    temp_33[lst_octant[k+1]]=temp_33[lst_octant[k+1]]+1
                if(lst_octant[k]==4):
                    temp_4[lst_octant[k+1]]=temp_4[lst_octant[k+1]]+1
                if(lst_octant[k]==-4):
                    temp_44[lst_octant[k+1]]=temp_44[lst_octant[k+1]]+1
            lst_tran_count_temp.append(temp_1)
            lst_tran_count_temp.append(temp_11)
            lst_tran_count_temp.append(temp_2)
            lst_tran_count_temp.append(temp_22)
            lst_tran_count_temp.append(temp_3)
            lst_tran_count_temp.append(temp_33)
            lst_tran_count_temp.append(temp_4)
            lst_tran_count_temp.append(temp_44)
            lst_tran_count.append(lst_tran_count_temp)
            lst_tran_count_temp=[]
            temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}

        lst_tran_count1=[]

        for i in range(t):

            lst_tran_count_temp=[]
            temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            for k in range(19886):

                if(lst_octant[k]==1):
                    temp_1[lst_octant[k+1]]=temp_1[lst_octant[k+1]]+1
                if(lst_octant[k]==-1):
                    temp_11[lst_octant[k+1]]=temp_11[lst_octant[k+1]]+1
                if(lst_octant[k]==2):
                    temp_2[lst_octant[k+1]]=temp_2[lst_octant[k+1]]+1
                if(lst_octant[k]==-2):
                    temp_22[lst_octant[k+1]]=temp_22[lst_octant[k+1]]+1
                if(lst_octant[k]==3):
                    temp_3[lst_octant[k+1]]=temp_3[lst_octant[k+1]]+1
                if(lst_octant[k]==-3):
                    temp_33[lst_octant[k+1]]=temp_33[lst_octant[k+1]]+1
                if(lst_octant[k]==4):
                    temp_4[lst_octant[k+1]]=temp_4[lst_octant[k+1]]+1
                if(lst_octant[k]==-4):
                    temp_44[lst_octant[k+1]]=temp_44[lst_octant[k+1]]+1
            lst_tran_count_temp.append(temp_1)
            lst_tran_count_temp.append(temp_11)
            lst_tran_count_temp.append(temp_2)
            lst_tran_count_temp.append(temp_22)
            lst_tran_count_temp.append(temp_3)
            lst_tran_count_temp.append(temp_33)
            lst_tran_count_temp.append(temp_4)
            lst_tran_count_temp.append(temp_44)
            lst_tran_count1.append(lst_tran_count_temp)
            lst_tran_count_temp=[]
            temp_1={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_11={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_2={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_22={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_3={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_33={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_4={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
            temp_44={1:0,-1:0,2:0,-2:0,3:0,-3:0,4:0,-4:0}
    except:
        print("there is some error in creating list of mod transition value")


    #### updating mod transition value
    try:
        lst_skelton_1=[[nan,"Overall Transition Count",nan,nan,nan,nan,nan,nan,nan,nan],[nan,nan,"To",nan,nan,nan,nan,nan,nan,nan],[nan,"count","+1","-1","+2","-2","+3","-3","+4","-4"],["From","+1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+4",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-4",nan,nan,nan,nan,nan,nan,nan,nan]]
        lst_skelton=[[nan,"Mod Transition count",nan,nan,nan,nan,nan,nan,nan,nan],[nan,nan,"To",nan,nan,nan,nan,nan,nan,nan],[nan,"count","+1","-1","+2","-2","+3","-3","+4","-4"],["From","+1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-1",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-2",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-3",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"+4",nan,nan,nan,nan,nan,nan,nan,nan],[nan,"-4",nan,nan,nan,nan,nan,nan,nan,nan]]
        i=0
        zz=0
        for row in sheet.iter_rows(min_row=zz, min_col=34, max_row=zz+10, max_col=43):
            j=0
            for cell in row:
                cell.value=lst_skelton_1[i][j]
                j=j+1
            i=i+1
        for h in range(1,t+1):
            i=0
            for row in sheet.iter_rows(min_row=13*h+zz, min_col=34, max_row=13*(h)+zz+10, max_col=43):
                j=0
                for cell in row:
                    cell.value=lst_skelton[i][j]
                    j=j+1
                i=i+1

        lst_tran_count_ans=[]
        for lsss in lst_tran_count:
            list_temp_1=[]
            for lsss1 in lsss:
                list_temp=[]
                for x in lsss1:
                    list_temp.append(lsss1[x])
                list_temp_1.append(list_temp)
            lst_tran_count_ans.append(list_temp_1)
        lst_tran_count_ans1=[]
        for lsss in lst_tran_count1:
            list_temp_1=[]
            for lsss1 in lsss:
                list_temp=[]
                for x in lsss1:
                    list_temp.append(lsss1[x])
                list_temp_1.append(list_temp)
            lst_tran_count_ans1.append(list_temp_1)

        for h in range(1,t+1):
            i=0
            for row in sheet.iter_rows(min_row=13*(h)+zz+3, min_col=36, max_row=13*(h)+zz+10, max_col=43):
                j=0
                for cell in row:
                    cell.value=lst_tran_count_ans[h-1][i][j]
                    j=j+1
                i=i+1
        i=0
        for h in range(1,t+1):
            for row in sheet.iter_rows(min_row=13*(h)+zz+1, min_col=35, max_row=13*(h)+zz+1, max_col=35):
                for cell in row:
                    cell.value=lst_hh[i]    
                i=i+1

        i=0
        for row in sheet.iter_rows(min_row=zz+3, min_col=36, max_row=zz+10, max_col=43):
            j=0
            for cell in row:
                cell.value=lst_tran_count_ans1[0][i][j]
                j=j+1
            i=i+1
    except:
        print("there is error in updating mod transition value in excel")

def octant_analysis(items,mod=5000):
    ##### imorting openpyxl and nan and loading workbook
    try:
        wb=openpyxl.load_workbook(r"C:\Users\DELL\OneDrive\Desktop\tt\tut07\input\{}".format(items))
    except:
        print("there is error in loading workbook check your file directory and import openpyxl")
        exit()
    sheet=wb.active
    t=0
    if(19888%mod==0):
        t=19888//mod
    else:
        t=(19888//mod)+1
    for i in range(t+3):
        for j in range(19):
            sheet.cell(row=i+2,column=j+13).border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
    for i in range(9):
        for j in range(3):
            sheet.cell(row=i+12,column=j+29).border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
    for i in range(9):
        for j in range(3):
            sheet.cell(row=i+2,column=j+45).border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))
    for h in range(t+1):
        for i in range(9):
            for j in range(9):
                sheet.cell(row=h*12+i+4,column=j+35).border=Border(left=Side(border_style='thin',color="000000"),right=Side(border_style='thin',color="000000"),top=Side(border_style='thin',color="000000"),bottom=Side(border_style='thin',color="000000"))

    ##### saving workbook file
    tut5(wb,mod)
    tut2(wb,mod)
    tut4(wb,mod)
    try:
        file_name=items.split(".")[0]
        file_name=file_name+'.'
        file_name=file_name+items.split(".")[1]
        wb.save(r'C:\Users\DELL\OneDrive\Desktop\tt\tut07\output\{} cm_vel_octant_analysis_mod_{}.xlsx'.format(file_name,mod))
    except:
        print("there is error in saving excel file 6666")

for i in range(1):
    octant_analysis(all_filesname[i],mod)
        
#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
